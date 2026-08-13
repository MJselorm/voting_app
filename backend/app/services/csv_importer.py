from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from email_validator import EmailNotValidError, validate_email

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.student import Student

REQUIRED = ("student_id", "full_name", "email", "department", "level", "class", "status")
HEADER_MAP = {"student id": "student_id", "student_id": "student_id", "full name": "full_name", "full_name": "full_name", "email": "email", "department": "department", "level": "level", "class": "class", "status": "status"}


@dataclass
class Preview:
    records: list[dict]
    errors: list[dict]
    total_rows: int
    new_count: int
    existing_count: int


def _header(value: str) -> str:
    return HEADER_MAP.get(re.sub(r"[_\s]+", " ", (value or "").strip().lower()), "")


def parse_upload(content: bytes, filename: str) -> tuple[list[dict[str, str]], list[dict]]:
    try:
        if filename.lower().endswith(".csv"):
            rows = list(csv.reader(io.StringIO(content.decode("utf-8-sig"))))
        elif filename.lower().endswith(".xlsx"):
            from openpyxl import load_workbook
            sheet = load_workbook(io.BytesIO(content), read_only=True, data_only=True).active
            rows = [["" if cell is None else str(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
        else:
            return [], [{"row": 0, "student_id": None, "reason": "Only CSV and Excel (.xlsx) files are supported."}]
    except Exception:
        return [], [{"row": 0, "student_id": None, "reason": "The uploaded file could not be read."}]
    if not rows:
        return [], [{"row": 0, "student_id": None, "reason": "The uploaded file is empty."}]
    headers = [_header(v) for v in rows[0]]
    missing = [x for x in REQUIRED if x not in headers]
    if missing:
        return [], [{"row": 1, "student_id": None, "reason": f"Missing required column(s): {', '.join(missing)}"}]
    records = []
    for number, cells in enumerate(rows[1:], 2):
        row = {headers[i]: (cells[i].strip() if i < len(cells) else "") for i in range(len(headers)) if headers[i]}
        row["email"] = row.get("email", "").lower()
        row["status"] = row.get("status", "").upper()
        row["row"] = number
        records.append(row)
    return records, []


async def preview_upload(content: bytes, filename: str, db: AsyncSession) -> Preview:
    records, errors = parse_upload(content, filename)
    if errors and not records:
        return Preview([], errors, 0, 0, 0)
    seen_ids, seen_emails = set(), set()
    existing = {s.student_id: s for s in (await db.execute(select(Student))).scalars()}
    emails = {s.email.lower(): s.student_id for s in existing.values() if s.email}
    new_count = existing_count = 0
    for record in records:
        reasons = []
        for field in REQUIRED:
            if not record.get(field): reasons.append(f"Missing required {field}.")
        if record.get("email"):
            try: validate_email(record["email"], check_deliverability=False)
            except EmailNotValidError: reasons.append("Invalid email address.")
        if record.get("status") not in {"ACTIVE", "INACTIVE"}: reasons.append("Status must be ACTIVE or INACTIVE.")
        sid, email = record.get("student_id", ""), record.get("email", "")
        if sid and sid.lower() in seen_ids: reasons.append("Duplicate Student ID inside uploaded file.")
        if email and email in seen_emails: reasons.append("Duplicate email inside uploaded file.")
        seen_ids.add(sid.lower()); seen_emails.add(email)
        if email in emails and emails[email] != sid: reasons.append("Email belongs to a different existing student.")
        if reasons:
            record["validation_status"] = "INVALID" if not any("Duplicate" in x for x in reasons) else "DUPLICATE"
            errors.append({"row": record["row"], "student_id": sid or None, "reason": " ".join(reasons)})
        elif sid in existing:
            record["validation_status"] = "EXISTING"; existing_count += 1
        else:
            record["validation_status"] = "NEW"; new_count += 1
    return Preview(records, errors, len(records), new_count, existing_count)


async def apply_preview(records: list[dict], behavior: str, db: AsyncSession) -> tuple[int, int, int]:
    added = updated = skipped = 0
    for rec in records:
        if rec.get("validation_status") not in {"NEW", "EXISTING"}: skipped += 1; continue
        current = (await db.execute(select(Student).where(Student.student_id == rec["student_id"]))).scalar_one_or_none()
        if current and behavior == "skip": skipped += 1; continue
        if current:
            for key in ("full_name", "email", "department", "level", "status"): setattr(current, key, rec[key])
            current.class_ = rec["class"]; updated += 1
        else:
            db.add(Student(student_id=rec["student_id"], full_name=rec["full_name"], email=rec["email"], department=rec["department"], level=rec["level"], class_=rec["class"], status=rec["status"])); added += 1
    await db.flush()
    return added, updated, skipped
